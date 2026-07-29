"""
知识库管理模块

封装 RAG 知识库的完整生命周期管理，包括文档上传解析、文本清洗分段、
向量入库、检索策略及后台管理接口。
"""

import logging
import os
from pathlib import Path
from typing import Any, List

from config.settings import rag_config, app_config
from core.rag.vector_store import VectorStore, chunk_text, clean_text
from core.rag.embedding_client import LocalEmbeddingClient

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 支持的文档解析器
# ---------------------------------------------------------------------------

class DocumentParser:
    """
    文档解析器：根据文件扩展名选择相应的解析策略
    """

    @staticmethod
    def parse(file_path: str | Path) -> str:
        """
        解析文档为纯文本

        Args:
            file_path: 文件路径

        Returns:
            提取的纯文本内容

        Raises:
            ValueError: 不支持的文档格式
        """
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".txt":
            return DocumentParser._parse_txt(path)
        elif suffix == ".md":
            return DocumentParser._parse_md(path)
        elif suffix == ".pdf":
            return DocumentParser._parse_pdf(path)
        elif suffix in (".xls", ".xlsx"):
            return DocumentParser._parse_excel(path)
        elif suffix in (".csv", ".tsv"):
            return DocumentParser._parse_csv(path)
        else:
            raise ValueError(f"不支持的文档格式: {suffix}")

    @staticmethod
    def _parse_txt(path: Path) -> str:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()

    @staticmethod
    def _parse_md(path: Path) -> str:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()

    @staticmethod
    def _parse_pdf(path: Path) -> str:
        try:
            import PyPDF2
        except ImportError:
            raise ImportError("解析 PDF 需要安装 PyPDF2: pip install PyPDF2")
        text_parts: List[str] = []
        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
        return "\n".join(text_parts)

    @staticmethod
    def _parse_excel(path: Path) -> str:
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "解析 Excel 需要安装 openpyxl: pip install openpyxl"
            )
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        text_parts: List[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: List[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("\t".join(cells))
            text_parts.append(f"【Sheet: {sheet_name}】\n" + "\n".join(rows))
        return "\n\n".join(text_parts)

    @staticmethod
    def _parse_csv(path: Path) -> str:
        import csv
        text_parts: List[str] = []
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                text_parts.append(", ".join(row))
        return "\n".join(text_parts)


# ---------------------------------------------------------------------------
# 知识库管理类
# ---------------------------------------------------------------------------

class KnowledgeBase:
    """
    知识库管理器

    封装了文档上传解析、自动清洗分段、向量入库、检索和管理接口。

    用法:
        kb = KnowledgeBase()
        kb.add_document("/path/to/report.pdf", metadata={"biz_line": "电商"})
        results = kb.search("上月销售额趋势")
        kb.delete_document("doc-id-xxx")
        kb.rebuild_index()
    """

    def __init__(
        self,
        collection_name: str = "data_analysis_kb",
        embedding_client: LocalEmbeddingClient | None = None,
        vector_store: VectorStore | None = None,
    ) -> None:
        """
        初始化知识库

        Args:
            collection_name: ChromaDB 集合名称
            embedding_client: Embedding 客户端
            vector_store: 向量存储实例（如提供则忽略 collection_name）
        """
        self._embedding_client = embedding_client or LocalEmbeddingClient()
        self._vector_store = vector_store or VectorStore(
            collection_name=collection_name,
            embedding_client=self._embedding_client,
        )
        self._collection_name = self._vector_store.collection_name
        logger.info("知识库初始化完成: collection=%s", self._collection_name)

    # ------------------------------------------------------------------
    # 属性
    # ------------------------------------------------------------------

    @property
    def collection_name(self) -> str:
        return self._collection_name

    @property
    def document_count(self) -> int:
        """返回知识库中的文档切片数量"""
        return self._vector_store.count

    @property
    def vector_store(self) -> VectorStore:
        """获取底层向量存储实例"""
        return self._vector_store

    # ------------------------------------------------------------------
    # 文档入库
    # ------------------------------------------------------------------

    def add_text(
        self,
        text: str,
        metadata: dict[str, Any] | None = None,
    ) -> int:
        """
        直接添加文本到知识库

        Args:
            text: 文本内容
            metadata: 元数据

        Returns:
            入库的切片数量
        """
        documents = [{"text": text, "metadata": metadata or {}}]
        return self._vector_store.add_documents(documents)

    def add_document(
        self,
        file_path: str | Path,
        metadata: dict[str, Any] | None = None,
    ) -> int:
        """
        上传并解析文档，然后入库

        Args:
            file_path: 文档路径
            metadata: 元数据（如 biz_line, metric_type, doc_time 等）

        Returns:
            入库的切片数量
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        logger.info("开始解析文档: %s", file_path)
        text = DocumentParser.parse(file_path)
        logger.info("文档解析完成: %s（%d 字符）", file_path, len(text))

        # 补充文件来源元数据
        meta = metadata or {}
        meta.setdefault("source", str(path.resolve()))
        meta.setdefault("filename", path.name)
        meta.setdefault("file_size", path.stat().st_size)

        return self.add_text(text, metadata=meta)

    def add_documents_batch(
        self,
        documents: List[dict[str, Any]],
    ) -> int:
        """
        批量添加文档（字典列表）

        每个文档支持两种格式:
        - {"text": "...", "metadata": {...}} —— 直接文本
        - {"file_path": "...", "metadata": {...}} —— 文件路径

        Args:
            documents: 文档列表

        Returns:
            总入库切片数量
        """
        total_chunks = 0
        for doc in documents:
            if "file_path" in doc:
                total_chunks += self.add_document(
                    doc["file_path"],
                    metadata=doc.get("metadata"),
                )
            elif "text" in doc:
                total_chunks += self.add_text(
                    doc["text"],
                    metadata=doc.get("metadata"),
                )
            else:
                logger.warning("文档格式无效，跳过: %s", doc)
        logger.info("批量入库完成，共 %d 切片", total_chunks)
        return total_chunks

    # ------------------------------------------------------------------
    # 检索
    # ------------------------------------------------------------------

    def search(
        self,
        query: str,
        top_k: int | None = None,
        similarity_threshold: float | None = None,
        filter_conditions: dict[str, Any] | None = None,
    ) -> List[dict[str, Any]]:
        """
        知识库检索

        Args:
            query: 查询文本
            top_k: 返回 Top-K 条结果
            similarity_threshold: 相似度阈值
            filter_conditions: 元数据过滤条件

        Returns:
            检索结果列表
        """
        return self._vector_store.search(
            query=query,
            top_k=top_k,
            similarity_threshold=similarity_threshold,
            filter_conditions=filter_conditions,
        )

    def search_with_sources(
        self,
        query: str,
        top_k: int | None = None,
        similarity_threshold: float | None = None,
        filter_conditions: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """
        检索结果并附带来源信息（供前端展示）

        Returns:
            {
                "query": "...",
                "results": [...],
                "total": 3,
                "collection": "data_analysis_kb",
            }
        """
        results = self.search(
            query=query,
            top_k=top_k,
            similarity_threshold=similarity_threshold,
            filter_conditions=filter_conditions,
        )
        return {
            "query": query,
            "results": results,
            "total": len(results),
            "collection": self._collection_name,
        }

    # ------------------------------------------------------------------
    # 管理接口
    # ------------------------------------------------------------------

    def list_documents(self, limit: int = 100) -> List[dict[str, Any]]:
        """
        列出知识库中的所有文档切片

        Args:
            limit: 最大返回条数

        Returns:
            文档列表
        """
        return self._vector_store.get_all_documents(limit=limit)

    def delete_document(self, doc_id: str) -> None:
        """
        删除指定文档切片（ChromaDB 原生不支持按 id 单独删除集合内文档需通过 delete）
        """
        try:
            self._vector_store._collection.delete(ids=[doc_id])
            logger.info("已删除文档切片: %s", doc_id)
        except Exception as e:
            logger.error("删除文档切片失败: %s", e)
            raise

    def delete_by_metadata(self, filter_conditions: dict[str, Any]) -> int:
        """
        根据元数据条件删除文档

        Args:
            filter_conditions: 过滤条件，e.g. {"biz_line": "电商"}

        Returns:
            删除的文档数（ChromaDB 未直接返回数量，仅记录日志）
        """
        try:
            self._vector_store._collection.delete(
                where=filter_conditions
            )
            logger.info(
                "已根据元数据条件删除文档: %s", filter_conditions
            )
            # 返回 -1 表示无法精确统计
            return -1
        except Exception as e:
            logger.error("根据元数据删除文档失败: %s", e)
            raise

    def rebuild_index(
        self,
        documents: List[dict[str, Any]] | None = None,
    ) -> int:
        """
        重建索引（清空集合后重新入库）

        Args:
            documents: 要重新入库的文档列表（如为 None 则仅清空）

        Returns:
            入库的切片数量
        """
        logger.warning("开始重建索引...")
        self._vector_store.reset_collection()

        if documents:
            total = self._vector_store.add_documents(documents)
            logger.info("索引重建完成，共 %d 切片", total)
            return total

        logger.info("索引已清空，未传入新文档")
        return 0

    def get_stats(self) -> dict[str, Any]:
        """
        获取知识库统计信息

        Returns:
            包含文档数量、集合名称和 Embedding 模型信息的字典
        """
        return {
            "collection_name": self._collection_name,
            "document_count": self.document_count,
            "embedding_model": self._embedding_client.model_name,
            "embedding_dim": self._embedding_client.embedding_dim,
            "chroma_db_dir": rag_config.chroma_db_dir,
            "top_k": rag_config.top_k,
            "similarity_threshold": rag_config.similarity_threshold,
        }
